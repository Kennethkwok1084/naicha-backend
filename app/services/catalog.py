"""商品目录服务层"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from fastapi import UploadFile
from sqlalchemy import and_, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.catalog import (
    Category,
    Product,
    ProductCategory,
    ProductSpecMapping,
    SpecGroup,
    SpecOption,
)
from app.schemas.catalog_admin import (
    CategoryCreateSchema,
    CategoryUpdateSchema,
    ProductCreateSchema,
    ProductImportResultSchema,
    ProductImportRowSchema,
    ProductUpdateSchema,
    SpecGroupCreateSchema,
    SpecGroupUpdateSchema,
    SpecOptionCreateSchema,
    SpecOptionUpdateSchema,
)


# ============ 分类服务 ============
async def get_categories(
    session: AsyncSession,
    skip: int = 0,
    limit: int = 100,
) -> tuple[list[Category], int]:
    """获取分类列表"""
    stmt = (
        select(Category)
        .order_by(Category.sort_order, Category.category_id)
        .offset(skip)
        .limit(limit)
    )
    result = await session.execute(stmt)
    items = list(result.scalars().all())

    count_stmt = select(func.count()).select_from(Category)
    total = await session.scalar(count_stmt) or 0

    return items, total


async def get_category_by_id(session: AsyncSession, category_id: int) -> Category | None:
    """根据 ID 获取分类"""
    stmt = select(Category).where(Category.category_id == category_id)
    result = await session.execute(stmt)
    return result.scalar_one_or_none()


async def create_category(
    session: AsyncSession,
    data: CategoryCreateSchema,
) -> Category:
    """创建分类"""
    category = Category(
        name=data.name,
        sort_order=data.sort_order,
    )
    session.add(category)
    await session.flush()
    await session.refresh(category)
    return category


async def update_category(
    session: AsyncSession,
    category_id: int,
    data: CategoryUpdateSchema,
) -> Category | None:
    """更新分类"""
    category = await get_category_by_id(session, category_id)
    if not category:
        return None

    if data.name is not None:
        category.name = data.name
    if data.sort_order is not None:
        category.sort_order = data.sort_order

    await session.flush()
    await session.refresh(category)
    return category


async def delete_category(session: AsyncSession, category_id: int) -> bool:
    """删除分类"""
    category = await get_category_by_id(session, category_id)
    if not category:
        return False
    await session.delete(category)
    await session.flush()
    return True


# ============ 商品服务 ============
async def get_products(
    session: AsyncSession,
    skip: int = 0,
    limit: int = 100,
    category_id: int | None = None,
    status: str | None = None,
    inventory_status: str | None = None,
    search: str | None = None,
) -> tuple[list[Product], int]:
    """获取商品列表"""
    stmt = select(Product).options(selectinload(Product.categories))

    filters = []
    if category_id is not None:
        filters.append(Product.category_id == category_id)
    if status is not None:
        filters.append(Product.status == status)
    if inventory_status is not None:
        filters.append(Product.inventory_status == inventory_status)
    if search:
        filters.append(Product.name.ilike(f"%{search}%"))

    if filters:
        stmt = stmt.where(and_(*filters))

    stmt = stmt.order_by(Product.product_id.desc()).offset(skip).limit(limit)
    result = await session.execute(stmt)
    items = list(result.scalars().all())

    # 填充 category_ids 和 spec_group_ids
    for product in items:
        product.category_ids = [c.category_id for c in product.categories]  # type: ignore
        # 获取规格组映射
        spec_stmt = select(ProductSpecMapping.group_id).where(
            ProductSpecMapping.product_id == product.product_id
        )
        spec_result = await session.execute(spec_stmt)
        product.spec_group_ids = list(spec_result.scalars().all())  # type: ignore

    # 计算总数
    count_stmt = select(func.count()).select_from(Product)
    if filters:
        count_stmt = count_stmt.where(and_(*filters))
    total = await session.scalar(count_stmt) or 0

    return items, total


async def get_product_by_id(session: AsyncSession, product_id: int) -> Product | None:
    """根据 ID 获取商品"""
    stmt = (
        select(Product)
        .options(selectinload(Product.categories))
        .where(Product.product_id == product_id)
    )
    result = await session.execute(stmt)
    product = result.scalar_one_or_none()

    if product:
        product.category_ids = [c.category_id for c in product.categories]  # type: ignore
        # 获取规格组映射
        spec_stmt = select(ProductSpecMapping.group_id).where(
            ProductSpecMapping.product_id == product.product_id
        )
        spec_result = await session.execute(spec_stmt)
        product.spec_group_ids = list(spec_result.scalars().all())  # type: ignore

    return product


async def create_product(
    session: AsyncSession,
    data: ProductCreateSchema,
) -> Product:
    """创建商品"""
    # 验证分类存在性
    if data.category_id is not None:
        category = await get_category_by_id(session, data.category_id)
        if not category:
            raise ValueError(f"分类 ID {data.category_id} 不存在")
    
    # 验证关联分类存在性
    if data.category_ids:
        for cat_id in data.category_ids:
            category = await get_category_by_id(session, cat_id)
            if not category:
                raise ValueError(f"分类 ID {cat_id} 不存在")
    
    # 验证规格组存在性
    if data.spec_group_ids:
        for group_id in data.spec_group_ids:
            group = await get_spec_group_by_id(session, group_id)
            if not group:
                raise ValueError(f"规格组 ID {group_id} 不存在")
    
    product = Product(
        category_id=data.category_id,
        name=data.name,
        description=data.description,
        image_url=data.image_url,
        base_price=data.base_price,
        status=data.status,
        inventory_status=data.inventory_status,
        stock_quantity=data.stock_quantity,
    )
    session.add(product)
    await session.flush()

    # 处理分类关联
    if data.category_ids:
        for cat_id in data.category_ids:
            pc = ProductCategory(product_id=product.product_id, category_id=cat_id)
            session.add(pc)

    # 处理规格组关联
    if data.spec_group_ids:
        for group_id in data.spec_group_ids:
            mapping = ProductSpecMapping(product_id=product.product_id, group_id=group_id)
            session.add(mapping)

    await session.flush()
    await session.refresh(product, ["categories"])
    product.category_ids = [c.category_id for c in product.categories]  # type: ignore
    product.spec_group_ids = data.spec_group_ids or []  # type: ignore
    return product


async def update_product(
    session: AsyncSession,
    product_id: int,
    data: ProductUpdateSchema,
) -> Product | None:
    """更新商品"""
    product = await get_product_by_id(session, product_id)
    if not product:
        return None

    # 验证并更新基础字段
    if data.category_id is not None:
        # 验证分类存在性
        category = await get_category_by_id(session, data.category_id)
        if not category:
            raise ValueError(f"分类 ID {data.category_id} 不存在")
        product.category_id = data.category_id
    if data.name is not None:
        product.name = data.name
    if data.description is not None:
        product.description = data.description
    if data.image_url is not None:
        product.image_url = data.image_url
    if data.base_price is not None:
        product.base_price = data.base_price
    if data.status is not None:
        product.status = data.status
    if data.inventory_status is not None:
        product.inventory_status = data.inventory_status
    if data.stock_quantity is not None:
        product.stock_quantity = data.stock_quantity

    # 更新分类关联
    if data.category_ids is not None:
        # 验证分类存在性
        for cat_id in data.category_ids:
            category = await get_category_by_id(session, cat_id)
            if not category:
                raise ValueError(f"分类 ID {cat_id} 不存在")
        # 删除旧关联
        await session.execute(
            delete(ProductCategory).where(ProductCategory.product_id == product_id)
        )
        # 添加新关联
        for cat_id in data.category_ids:
            pc = ProductCategory(product_id=product_id, category_id=cat_id)
            session.add(pc)

    # 更新规格组关联
    if data.spec_group_ids is not None:
        # 验证规格组存在性
        for group_id in data.spec_group_ids:
            group = await get_spec_group_by_id(session, group_id)
            if not group:
                raise ValueError(f"规格组 ID {group_id} 不存在")
        # 删除旧关联
        await session.execute(
            delete(ProductSpecMapping).where(ProductSpecMapping.product_id == product_id)
        )
        # 添加新关联
        for group_id in data.spec_group_ids:
            mapping = ProductSpecMapping(product_id=product_id, group_id=group_id)
            session.add(mapping)

    await session.flush()
    await session.refresh(product, ["categories"])
    product.category_ids = [c.category_id for c in product.categories]  # type: ignore

    # 获取更新后的规格组
    spec_stmt = select(ProductSpecMapping.group_id).where(
        ProductSpecMapping.product_id == product_id
    )
    spec_result = await session.execute(spec_stmt)
    product.spec_group_ids = list(spec_result.scalars().all())  # type: ignore

    return product


async def delete_product(session: AsyncSession, product_id: int) -> bool:
    """删除商品"""
    product = await get_product_by_id(session, product_id)
    if not product:
        return False
    await session.delete(product)
    await session.flush()
    return True


async def batch_update_product_status(
    session: AsyncSession,
    product_ids: list[int],
    status: str | None = None,
    inventory_status: str | None = None,
) -> int:
    """批量更新商品状态"""
    updates: dict[str, Any] = {}
    if status is not None:
        updates["status"] = status
    if inventory_status is not None:
        updates["inventory_status"] = inventory_status

    if not updates:
        return 0

    stmt = select(Product).where(Product.product_id.in_(product_ids))
    result = await session.execute(stmt)
    products = list(result.scalars().all())

    for product in products:
        for key, value in updates.items():
            setattr(product, key, value)

    await session.flush()
    return len(products)


# ============ 规格组服务 ============
async def get_spec_groups(
    session: AsyncSession,
    skip: int = 0,
    limit: int = 100,
) -> tuple[list[SpecGroup], int]:
    """获取规格组列表"""
    stmt = (
        select(SpecGroup)
        .order_by(SpecGroup.sort_order, SpecGroup.group_id)
        .offset(skip)
        .limit(limit)
    )
    result = await session.execute(stmt)
    items = list(result.scalars().all())

    count_stmt = select(func.count()).select_from(SpecGroup)
    total = await session.scalar(count_stmt) or 0

    return items, total


async def get_spec_group_by_id(session: AsyncSession, group_id: int) -> SpecGroup | None:
    """根据 ID 获取规格组"""
    stmt = select(SpecGroup).where(SpecGroup.group_id == group_id)
    result = await session.execute(stmt)
    return result.scalar_one_or_none()


async def create_spec_group(
    session: AsyncSession,
    data: SpecGroupCreateSchema,
) -> SpecGroup:
    """创建规格组"""
    group = SpecGroup(
        name=data.name,
        sort_order=data.sort_order,
    )
    session.add(group)
    await session.flush()
    await session.refresh(group)
    return group


async def update_spec_group(
    session: AsyncSession,
    group_id: int,
    data: SpecGroupUpdateSchema,
) -> SpecGroup | None:
    """更新规格组"""
    group = await get_spec_group_by_id(session, group_id)
    if not group:
        return None

    if data.name is not None:
        group.name = data.name
    if data.sort_order is not None:
        group.sort_order = data.sort_order

    await session.flush()
    await session.refresh(group)
    return group


async def delete_spec_group(session: AsyncSession, group_id: int) -> bool:
    """删除规格组"""
    group = await get_spec_group_by_id(session, group_id)
    if not group:
        return False
    await session.delete(group)
    await session.flush()
    return True


# ============ 规格选项服务 ============
async def get_spec_options(
    session: AsyncSession,
    group_id: int | None = None,
    skip: int = 0,
    limit: int = 100,
) -> tuple[list[SpecOption], int]:
    """获取规格选项列表"""
    stmt = select(SpecOption)

    if group_id is not None:
        stmt = stmt.where(SpecOption.group_id == group_id)

    stmt = (
        stmt.order_by(SpecOption.group_id, SpecOption.sort_order, SpecOption.option_id)
        .offset(skip)
        .limit(limit)
    )
    result = await session.execute(stmt)
    items = list(result.scalars().all())

    count_stmt = select(func.count()).select_from(SpecOption)
    if group_id is not None:
        count_stmt = count_stmt.where(SpecOption.group_id == group_id)
    total = await session.scalar(count_stmt) or 0

    return items, total


async def get_spec_option_by_id(session: AsyncSession, option_id: int) -> SpecOption | None:
    """根据 ID 获取规格选项"""
    stmt = select(SpecOption).where(SpecOption.option_id == option_id)
    result = await session.execute(stmt)
    return result.scalar_one_or_none()


async def create_spec_option(
    session: AsyncSession,
    data: SpecOptionCreateSchema,
) -> SpecOption:
    """创建规格选项"""
    # 验证规格组存在性
    group = await get_spec_group_by_id(session, data.group_id)
    if not group:
        raise ValueError(f"规格组 ID {data.group_id} 不存在")
    
    option = SpecOption(
        group_id=data.group_id,
        name=data.name,
        price_modifier=data.price_modifier,
        inventory_status=data.inventory_status,
        sort_order=data.sort_order,
    )
    session.add(option)
    await session.flush()
    await session.refresh(option)
    return option


async def update_spec_option(
    session: AsyncSession,
    option_id: int,
    data: SpecOptionUpdateSchema,
) -> SpecOption | None:
    """更新规格选项"""
    option = await get_spec_option_by_id(session, option_id)
    if not option:
        return None

    if data.name is not None:
        option.name = data.name
    if data.price_modifier is not None:
        option.price_modifier = data.price_modifier
    if data.inventory_status is not None:
        option.inventory_status = data.inventory_status
    if data.sort_order is not None:
        option.sort_order = data.sort_order

    await session.flush()
    await session.refresh(option)
    return option


async def delete_spec_option(session: AsyncSession, option_id: int) -> bool:
    """删除规格选项"""
    option = await get_spec_option_by_id(session, option_id)
    if not option:
        return False
    await session.delete(option)
    await session.flush()
    return True


# ============ 批量导入服务 ============
# 导入文件限制配置
MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_IMPORT_ROWS = 5000  # 最多 5000 行
MAX_IMPORT_COLUMNS = 20  # 最多 20 列

# 允许的 MIME 类型
ALLOWED_IMPORT_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
    "text/csv",  # .csv
    "application/csv",
}


async def parse_product_import_file(file: UploadFile) -> list[ProductImportRowSchema]:
    """解析批量导入文件（Excel/CSV）

    安全限制：
    - 文件大小 <= 10MB
    - 行数 <= 5000
    - 列数 <= 20
    - MIME 类型校验
    """
    # 1. 验证 MIME 类型
    if file.content_type not in ALLOWED_IMPORT_MIME_TYPES:
        raise ValueError(
            f"不支持的文件类型: {file.content_type}。仅支持 Excel (.xlsx, .xls) 或 CSV 文件"
        )

    # 2. 读取并验证文件大小
    content = await file.read()
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f"文件过大: {len(content)} 字节，最大允许 {MAX_IMPORT_FILE_SIZE} 字节")

    # 3. 验证文件头（魔数）
    if file.filename and file.filename.endswith((".xlsx", ".xls")):
        # Excel 文件应该以 PK (ZIP 魔数) 开头
        if not content.startswith(b"PK"):
            raise ValueError("文件头验证失败：不是有效的 Excel 文件")
    elif file.filename and file.filename.endswith(".csv"):
        # CSV 应该是文本文件，尝试解码验证
        try:
            content.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise ValueError("文件编码验证失败：不是有效的 CSV 文件")

    rows: list[ProductImportRowSchema] = []

    try:
        # 支持 Excel
        if file.filename and file.filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            
            # 4. 验证列数
            if ws.max_column and ws.max_column > MAX_IMPORT_COLUMNS:
                raise ValueError(f"文件列数过多: {ws.max_column}，最大允许 {MAX_IMPORT_COLUMNS}")
            
            row_count = 0
            for idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):  # 跳过表头
                if not row or not row[0]:  # 跳过空行
                    continue
                row_count += 1
                # 5. 验证行数
                if row_count > MAX_IMPORT_ROWS:
                    raise ValueError(f"文件行数过多，最大允许 {MAX_IMPORT_ROWS} 行数据")
                rows.append(
                    ProductImportRowSchema(
                        row_number=idx,
                        category_name=row[0] if len(row) > 0 else None,
                        product_name=row[1] if len(row) > 1 else "",
                        description=row[2] if len(row) > 2 else None,
                        base_price=float(row[3]) if len(row) > 3 else 0.0,
                        status=row[4] if len(row) > 4 else "active",
                        inventory_status=row[5] if len(row) > 5 else "in_stock",
                        stock_quantity=int(row[6]) if len(row) > 6 else 0,
                        spec_groups=row[7] if len(row) > 7 else None,
                    )
                )
        # 支持 CSV
        elif file.filename and file.filename.endswith(".csv"):
            import csv

            content_str = content.decode("utf-8-sig")  # 处理 BOM
            reader = csv.reader(io.StringIO(content_str))
            next(reader, None)  # 跳过表头
            row_count = 0
            for idx, row in enumerate(reader, start=2):
                if not row or not row[0]:
                    continue
                row_count += 1
                if row_count > MAX_IMPORT_ROWS:
                    raise ValueError(f"文件行数过多，最大允许 {MAX_IMPORT_ROWS} 行数据")
                if len(row) > MAX_IMPORT_COLUMNS:
                    raise ValueError(f"CSV 列数过多: {len(row)}，最大允许 {MAX_IMPORT_COLUMNS}")
                rows.append(
                    ProductImportRowSchema(
                        row_number=idx,
                        category_name=row[0] if len(row) > 0 else None,
                        product_name=row[1] if len(row) > 1 else "",
                        description=row[2] if len(row) > 2 else None,
                        base_price=float(row[3]) if len(row) > 3 else 0.0,
                        status=row[4] if len(row) > 4 else "active",
                        inventory_status=row[5] if len(row) > 5 else "in_stock",
                        stock_quantity=int(row[6]) if len(row) > 6 else 0,
                        spec_groups=row[7] if len(row) > 7 else None,
                    )
                )
    except Exception as e:
        raise ValueError(f"文件解析失败: {e}")

    return rows


async def batch_import_products(
    session: AsyncSession,
    rows: list[ProductImportRowSchema],
) -> ProductImportResultSchema:
    """批量导入商品"""
    result = ProductImportResultSchema(success_count=0, error_count=0, errors=[])

    # 预加载所有分类和规格组（避免重复查询）
    category_cache: dict[str, int] = {}
    spec_group_cache: dict[str, int] = {}

    cat_stmt = select(Category)
    cat_result = await session.execute(cat_stmt)
    for cat in cat_result.scalars().all():
        category_cache[cat.name] = cat.category_id

    spec_stmt = select(SpecGroup)
    spec_result = await session.execute(spec_stmt)
    for group in spec_result.scalars().all():
        spec_group_cache[group.name] = group.group_id

    for row in rows:
        try:
            # 验证必填字段
            if not row.product_name:
                raise ValueError("商品名称不能为空")
            if row.base_price <= 0:
                raise ValueError("商品价格必须大于0")

            # 查找或创建分类
            category_id = None
            if row.category_name:
                if row.category_name in category_cache:
                    category_id = category_cache[row.category_name]
                else:
                    # 自动创建分类
                    new_cat = Category(name=row.category_name, sort_order=0)
                    session.add(new_cat)
                    await session.flush()
                    category_id = new_cat.category_id
                    category_cache[row.category_name] = category_id

            # 创建商品
            product = Product(
                category_id=category_id,
                name=row.product_name,
                description=row.description,
                base_price=row.base_price,
                status=row.status,
                inventory_status=row.inventory_status,
                stock_quantity=row.stock_quantity,
            )
            session.add(product)
            await session.flush()

            # 处理规格组关联
            if row.spec_groups:
                spec_names = [s.strip() for s in row.spec_groups.split(",") if s.strip()]
                for spec_name in spec_names:
                    if spec_name in spec_group_cache:
                        mapping = ProductSpecMapping(
                            product_id=product.product_id,
                            group_id=spec_group_cache[spec_name],
                        )
                        session.add(mapping)

            result.success_count += 1

        except Exception as e:
            result.error_count += 1
            result.errors.append(
                {
                    "row_number": row.row_number,
                    "error_message": str(e),
                }
            )

    await session.flush()
    return result
